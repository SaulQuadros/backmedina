"""Geração da planilha padronizada "Tabela" (formato Modelo.xlsx).

Reproduz o layout SOLOCAP: bloco de metadados (linhas 1-10) + cabeçalho das
colunas (linha 13) + dados (linha 14+). Serve como saída fiel ao .exe original.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook

from backmedina.model.schema import (
    ABA_TABELA,
    COLUNAS_TABELA,
    D_TO_SENSOR,
    UNIDADE_DEFLEXAO,
    DadosFWD,
)
from backmedina.model.units import UnidadeDeflexao, fator_conversao

# Colunas de deflexão da planilha: D1..D9 (mapa de sensores) + D10 (externo).
_COLUNAS_DEFLEXAO: tuple[str, ...] = (*D_TO_SENSOR, "D10")


def _tabela_em_001mm(dados: DadosFWD) -> pd.DataFrame:
    """Deflexões convertidas para 0,01 mm — a unidade que esta planilha declara.

    Sem isto, uma fonte em µm (CSV de bacias, KUAB) era gravada com os valores
    crus sob o rótulo `x10⁻² mm`: a planilha saía mentindo e todo o fluxo a
    jusante (índices, Rc, segmentação, CSV BackMeDiNa) errava por 10x.
    """
    fator = fator_conversao(dados.unidade_deflexao, UnidadeDeflexao.DMM_001)
    df: pd.DataFrame = dados.tabela
    if fator == 1.0:  # já está em 0,01 mm — preserva os valores originais
        return df
    df = df.copy()
    for col in _COLUNAS_DEFLEXAO:
        if col in df.columns:
            # round(3): mata o ruído de ponto flutuante (706 × 0,1 =
            # 70,60000000000001) sem perder precisão real — a fonte em µm é
            # inteira, então 1 casa decimal já basta em 0,01 mm.
            df[col] = (pd.to_numeric(df[col], errors="coerce") * fator).round(3)
    return df


def _linhas_metadados(dados: DadosFWD) -> list[tuple[str, str]]:
    m = dados.metadados
    return [
        ("DATA", m.data),
        ("RELATÓRIO N°", m.relatorio),
        ("OS Nº", m.os_numero),
        ("CLIENTE", m.cliente),
        ("OBRA/TRECHO", m.obra_trecho),
        ("PISTA", m.pista),
        ("SENTIDO", m.sentido),
        ("UNIDADE DAS LEITURAS", f"UNIDADE DAS LEITURAS ({UNIDADE_DEFLEXAO})"),
    ]


def montar_workbook(dados: DadosFWD) -> Workbook:
    """Monta o Workbook openpyxl com a aba "Tabela" padronizada."""
    wb = Workbook()
    ws = wb.active
    ws.title = ABA_TABELA

    ws.cell(1, 1, "LEVANTAMENTO DEFLECTOMÉTRICO (FWD) — Tabela extraída do PDF")
    for i, (chave, valor) in enumerate(_linhas_metadados(dados), start=3):
        ws.cell(i, 1, chave)
        ws.cell(i, 2, valor)

    # Cabeçalho das colunas na linha 13.
    for c, nome in enumerate(COLUNAS_TABELA, start=1):
        ws.cell(13, c, nome)

    # Dados a partir da linha 14, respeitando a ordem de COLUNAS_TABELA.
    # Deflexões normalizadas para 0,01 mm (o rótulo da linha 10 promete isso).
    df: pd.DataFrame = _tabela_em_001mm(dados)
    for r, (_, row) in enumerate(df.iterrows(), start=14):
        for c, nome in enumerate(COLUNAS_TABELA, start=1):
            valor = row[nome] if nome in df.columns else ""
            if pd.isna(valor):
                valor = ""
            ws.cell(r, c, valor)
    return wb


def exportar_xlsx_bytes(dados: DadosFWD) -> bytes:
    """Serializa a planilha padronizada em bytes (para download no Streamlit)."""
    wb = montar_workbook(dados)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
