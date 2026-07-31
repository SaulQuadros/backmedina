"""Rascunho .xlsx a partir de um PDF que nenhum leitor reconheceu.

Quando o layout do relatório não é conhecido, o App não tem como mapear as
colunas — e um palpite errado é pior que nenhum. Este rascunho entrega o que o
App realmente sabe, para o usuário montar a planilha à mão e reimportar:

- aba **Tabela**: o esqueleto do layout SOLOCAP (metadados + cabeçalho das
  colunas), vazio, pronto para receber os dados;
- aba **Texto_do_PDF**: o texto extraído, uma linha por linha do PDF, já
  quebrado em colunas pelos espaços — é daí que se copia para a aba Tabela.

**A unidade das leituras sai em branco de propósito.** Preencher com um padrão
é como nasce o erro de 10x: a planilha passa a declarar 0,01 mm sobre valores em
µm. Em branco, o leitor entende "não declarado" e a página de importação avisa e
deixa escolher antes de qualquer cálculo.
"""

from __future__ import annotations

import io

from openpyxl import Workbook

from backmedina.model.schema import (
    ABA_TABELA,
    COLUNAS_TABELA,
    LINHA_HEADER_COLUNAS,
    METADADOS_CHAVES,
)

_TITULO = (
    "RASCUNHO — preencha os dados e a linha UNIDADE DAS LEITURAS antes de reimportar"
)
_MAX_COLUNAS_TEXTO = 40  # trava contra linhas absurdamente largas


def rascunho_xlsx_bytes(texto: str, origem: str = "") -> bytes:
    """Monta o rascunho (bytes .xlsx) a partir do texto extraído do PDF."""
    wb = Workbook()

    ws = wb.active
    ws.title = ABA_TABELA
    ws.cell(1, 1, _TITULO)
    if origem:
        ws.cell(1, 2, f"origem: {origem}")
    # Metadados: chaves preenchidas, valores em branco — inclusive a unidade.
    for i, chave in enumerate(METADADOS_CHAVES, start=3):
        ws.cell(i, 1, chave)
    for c, nome in enumerate(COLUNAS_TABELA, start=1):
        ws.cell(LINHA_HEADER_COLUNAS, c, nome)

    wt = wb.create_sheet("Texto_do_PDF")
    wt.cell(1, 1, "Texto extraído do PDF — copie para a aba Tabela na ordem correta")
    linha_saida = 3
    for linha in texto.splitlines():
        if not linha.strip():
            continue
        for c, campo in enumerate(linha.split()[:_MAX_COLUNAS_TEXTO], start=1):
            wt.cell(linha_saida, c, campo)
        linha_saida += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
