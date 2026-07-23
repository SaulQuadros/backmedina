"""Leitura da planilha SOLOCAP (aba "Tabela") -> DadosFWD.

Layout esperado (confirmado no arquivo real da UFJF):
  linhas 1-10  : bloco de metadados (chave em col A, valor em col B)
  linha  13    : cabeçalho das colunas de dados
  linha  14+   : dados (uma linha por estação/estaca)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from backmedina.io.br_numbers import parse_br_series
from backmedina.model.schema import (
    ABA_TABELA,
    COLUNAS_NUMERICAS,
    COLUNAS_TABELA,
    LINHA_HEADER_COLUNAS,
    LINHA_PRIMEIRO_DADO,
    MetadadosLevantamento,
    DadosFWD,
)
from backmedina.model.units import detectar_unidade

# Mapa chave-de-metadado (na planilha) -> atributo de MetadadosLevantamento.
_META_MAP = {
    "DATA": "data",
    "RELATÓRIO N°": "relatorio",
    "OS Nº": "os_numero",
    "CLIENTE": "cliente",
    "OBRA/TRECHO": "obra_trecho",
    "PISTA": "pista",
    "SENTIDO": "sentido",
    "UNIDADE DAS LEITURAS": "unidade",
}


def _ler_metadados(ws) -> MetadadosLevantamento:
    meta = MetadadosLevantamento()
    for r in range(1, LINHA_HEADER_COLUNAS):
        chave = ws.cell(r, 1).value
        valor = ws.cell(r, 2).value
        if chave is None:
            continue
        attr = _META_MAP.get(str(chave).strip())
        if attr and valor is not None:
            setattr(meta, attr, str(valor).strip())
    return meta


def ler_solocap_xlsx(caminho: str | Path) -> DadosFWD:
    """Lê uma planilha SOLOCAP e devolve :class:`DadosFWD` com tabela numérica."""
    caminho = Path(caminho)
    # read_only=False: os arquivos são pequenos e o acesso aleatório por célula
    # (ws.cell) é ordens de magnitude mais rápido fora do modo read_only.
    wb = openpyxl.load_workbook(caminho, data_only=True)
    aba = ABA_TABELA if ABA_TABELA in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]

    metadados = _ler_metadados(ws)

    # Lê as linhas de dados como texto cru.
    linhas: list[list[object]] = []
    for r in range(LINHA_PRIMEIRO_DADO, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, len(COLUNAS_TABELA) + 1)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        linhas.append(row)
    wb.close()

    df = pd.DataFrame(linhas, columns=list(COLUNAS_TABELA))

    avisos: list[str] = []
    for col in COLUNAS_NUMERICAS:
        df[col] = parse_br_series(df[col])
    # Obs é textual.
    df["Data e Hora"] = df["Data e Hora"].map(
        lambda v: "" if v is None else str(v).strip()
    )
    df["Obs"] = df["Obs"].map(lambda v: "" if v is None else str(v).strip())

    if df.empty:
        avisos.append("Nenhuma linha de dados encontrada na planilha.")

    unidade = detectar_unidade(metadados.unidade)

    return DadosFWD(
        metadados=metadados,
        tabela=df,
        origem=caminho.name,
        avisos=avisos,
        unidade_deflexao=unidade,
    )
