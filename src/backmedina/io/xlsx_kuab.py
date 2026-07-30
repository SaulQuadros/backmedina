"""Leitura da planilha de levantamento FWD do equipamento **KUAB** -> DadosFWD.

Formato distinto do SOLOCAP (referência: `z_docs/lwd/machado/`):

    linha 1..N   bloco de metadados — chave na col A, valor na 1ª célula à direita
                 (KUAB/cliente, Operador, Local, Clima, Data, Raio (cm),
                  Número Sensores, Distância Sensores (cm), Posição Sensor, …)
    linha H      cabeçalho: Distancia_m, Estaca_Numero, Load_kgf,
                 D0_um … D8_um, Air_C, Pave_C, Emod_MPa, Time
    linha H+1…   dados

Diferenças que este leitor resolve:

- **Nomenclatura dos geofones desloca em 1**: `D0_um` é o sensor central, que no
  mapa de Rocha/BackMeDiNa é `D1`. Logo `D0_um→D1 … D8_um→D9`. Não há 10º
  geofone (210 cm), então `D10` fica vazio.
- **Unidade µm** (declarada no próprio nome das colunas), contra 0,01 mm do
  SOLOCAP. Vai em `DadosFWD.unidade_deflexao` e o núcleo normaliza a partir daí.
- **Data só no cabeçalho** (uma para todo o levantamento); a linha traz apenas
  `Time`. As duas são combinadas em "Data e Hora".
- `Raio (cm)` do cabeçalho é o **raio da placa em mm** (150 mm = 15 cm), mal
  rotulado — não confundir com a coluna `Raio` do SOLOCAP, que é o raio de
  curvatura por estação. A coluna `Raio` sai vazia.
- `Estaca_Numero` é preservada (o SOLOCAP não tem, e o CSV BackMeDiNa usa).
- `Emod_MPa` (módulo retroanalisado pelo equipamento) é mantido como coluna
  extra, apenas informativa: retroanálise é do BackMeDiNa, não deste App.

As distâncias radiais declaradas na planilha são **conferidas** contra
`SENSOR_OFFSETS_CM`; divergência vira aviso, porque invalidaria o mapeamento.
"""

from __future__ import annotations

from datetime import datetime, time
from pathlib import Path

import openpyxl
import pandas as pd

from backmedina.io.br_numbers import parse_br_series
from backmedina.model.schema import (
    COLUNAS_NUMERICAS,
    COLUNAS_TABELA,
    SENSOR_OFFSETS_CM,
    MetadadosLevantamento,
    DadosFWD,
)
from backmedina.model.units import UnidadeDeflexao, detectar_unidade_explicita

# Colunas do KUAB -> colunas canônicas (chaves já normalizadas por _norm).
_MAPA_COLUNAS: dict[str, str] = {
    "distancia_m": "Metros",
    "load_kgf": "Target Load (Kgf)",
    "air_c": "Temp Ar (°C)",
    "pave_c": "Temp Pav (°C)",
    "estaca_numero": "Estaca – Número",
    "emod_mpa": "Emod_MPa",
}

# D0_um..D8_um -> D1..D9 (o central do KUAB é o D1 do mapa de sensores).
_MAPA_SENSORES: dict[str, str] = {
    f"d{i}_um": f"D{i + 1}" for i in range(len(SENSOR_OFFSETS_CM))
}

_LINHAS_META_MAX = 30  # o cabeçalho de dados aparece bem antes disto


def _e_chave_distancias(chave: str) -> bool:
    """Reconhece 'Distância Sensores (cm)' com ou sem acento/variação de grafia."""
    return chave.startswith("dist") and "sensor" in chave


def _norm(v) -> str:
    """Normaliza um rótulo para comparação: minúsculas, sem espaços nas pontas."""
    return str(v).strip().lower() if v is not None else ""


def _linha_cabecalho(ws) -> int | None:
    """Linha (1-indexada) do cabeçalho de dados: a que contém `D0_um`."""
    for r in range(1, min(ws.max_row, _LINHAS_META_MAX) + 1):
        rotulos = {_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if "d0_um" in rotulos:
            return r
    return None


def e_planilha_kuab(caminho: str | Path) -> bool:
    """Reconhece a planilha pelo cabeçalho de dados (não abre exceção)."""
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=False)
    except Exception:  # noqa: BLE001 — arquivo ilegível não é KUAB
        return False
    try:
        return any(_linha_cabecalho(ws) is not None for ws in wb.worksheets)
    finally:
        wb.close()


def _valor_a_direita(ws, r: int, ate: int) -> object:
    """1ª célula não vazia à direita da chave (o KUAB usa col C, não B)."""
    for c in range(2, ate + 1):
        v = ws.cell(r, c).value
        if v is not None and str(v).strip() != "":
            return v
    return None


def _ler_metadados(ws, linha_header: int) -> tuple[MetadadosLevantamento, str, list[str]]:
    """Metadados + data do levantamento + avisos das distâncias dos sensores."""
    meta = MetadadosLevantamento()
    data = ""
    avisos: list[str] = []

    for r in range(1, linha_header):
        chave = _norm(ws.cell(r, 1).value)
        if not chave:
            continue
        valor = _valor_a_direita(ws, r, ws.max_column)

        if chave == "kuab":  # a 1ª linha traz o equipamento e o cliente
            if valor is not None:
                meta.cliente = str(valor).strip()
        elif chave == "local":
            if valor is not None:
                meta.obra_trecho = str(valor).strip()
        elif chave == "data":
            data = _formatar_data(valor)
            meta.data = data
        elif _e_chave_distancias(chave):
            avisos += _conferir_offsets(ws, r)

    return meta, data, avisos


def _formatar_data(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


def _conferir_offsets(ws, r: int) -> list[str]:
    """Compara as distâncias radiais declaradas com o mapa de sensores."""
    lidos: list[float] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(r, c).value
        if isinstance(v, (int, float)):
            lidos.append(float(v))
    esperado = [float(o) for o in SENSOR_OFFSETS_CM]
    if lidos and lidos != esperado[: len(lidos)]:
        return [
            "Distâncias dos sensores na planilha "
            f"({', '.join(f'{v:g}' for v in lidos)} cm) diferem do mapa esperado "
            f"({', '.join(f'{v:g}' for v in esperado)} cm) — o mapeamento "
            "D0_um→d0 … D8_um→d180 pode não valer para este levantamento."
        ]
    return []


def _hora(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, time):
        return valor.strftime("%H:%M:%S")
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M:%S")
    return str(valor).strip()


def ler_kuab_xlsx(caminho: str | Path) -> DadosFWD:
    """Lê uma planilha FWD do KUAB e devolve :class:`DadosFWD` (deflexões em µm)."""
    caminho = Path(caminho)
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = next(
        (w for w in wb.worksheets if _linha_cabecalho(w) is not None), wb.worksheets[0]
    )
    linha_header = _linha_cabecalho(ws)
    if linha_header is None:
        wb.close()
        raise ValueError(
            "Planilha KUAB sem a coluna 'D0_um' — cabeçalho de dados não encontrado."
        )

    metadados, data_levantamento, avisos = _ler_metadados(ws, linha_header)

    # Cabeçalho original -> coluna canônica (as não mapeadas são descartadas).
    origem_col: dict[int, str] = {}
    rotulos: list[str] = []
    for c in range(1, ws.max_column + 1):
        rot = _norm(ws.cell(linha_header, c).value)
        if not rot:
            continue
        rotulos.append(rot)
        destino = _MAPA_SENSORES.get(rot) or _MAPA_COLUNAS.get(rot)
        if destino:
            origem_col[c] = destino
        elif rot == "time":
            origem_col[c] = "_Time"

    registros: list[dict] = []
    for r in range(linha_header + 1, ws.max_row + 1):
        bruto = {nome: ws.cell(r, c).value for c, nome in origem_col.items()}
        if all(v is None or str(v).strip() == "" for v in bruto.values()):
            continue
        hora = _hora(bruto.pop("_Time", None))
        bruto["Data e Hora"] = f"{data_levantamento} {hora}".strip()
        registros.append(bruto)

    wb.close()

    # Mesmo formato de tabela do SOLOCAP (+ extras), para o restante do fluxo
    # tratar as duas origens da mesma forma.
    extras = ["Estaca – Número", "Emod_MPa"]
    df = pd.DataFrame(registros, columns=[*COLUNAS_TABELA, *extras])
    for col in [*COLUNAS_NUMERICAS, *extras]:
        df[col] = parse_br_series(df[col])
    df["Obs"] = ""

    if df.empty:
        avisos.append("Nenhuma linha de dados encontrada na planilha KUAB.")

    # A unidade é DECLARADA nos nomes das colunas (`D0_um`): não presumir.
    unidade = detectar_unidade_explicita(" ".join(rotulos)) or UnidadeDeflexao.DMM_001
    metadados.unidade = unidade.rotulo

    return DadosFWD(
        metadados=metadados,
        tabela=df,
        origem=caminho.name,
        avisos=avisos,
        unidade_deflexao=unidade,
    )
