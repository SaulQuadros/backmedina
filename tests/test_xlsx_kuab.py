"""Leitor da planilha FWD do equipamento KUAB (layout distinto do SOLOCAP)."""

from pathlib import Path

import openpyxl
import pytest

from backmedina.io.loader import carregar
from backmedina.io.xlsx_kuab import e_planilha_kuab, ler_kuab_xlsx
from backmedina.model.schema import COLUNAS_TABELA, SENSOR_OFFSETS_CM
from backmedina.model.units import UnidadeDeflexao

RAIZ = Path(__file__).resolve().parents[1]
KUAB = RAIZ / "z_docs" / "lwd" / "machado" / "00_tab_original_FWD_machado_2019.xlsx"
SOLOCAP = RAIZ / "z_docs" / "lwd" / "solocap" / "2-UFJF-VIA_LOCAL_FX1-FWD.xlsx"

# Primeira estação do arquivo real (linha 15 da aba "Dados").
_PRIMEIRA = dict(
    metros=0.0, carga=4097.0, estaca=0.0,
    d=[410, 294, 212, 126, 92, 61, 45, 34, 26],  # D0_um..D8_um
)


def _kuab_sintetico(tmp_path, offsets=SENSOR_OFFSETS_CM, linhas=3):
    """Planilha KUAB mínima — evita depender do arquivo real nos testes básicos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.cell(1, 1, "KUAB"); ws.cell(1, 3, "CLIENTE X")
    ws.cell(2, 1, "Local"); ws.cell(2, 3, "TRECHO Y")
    ws.cell(3, 1, "Data"); ws.cell(3, 3, "10/03/2018")
    ws.cell(4, 1, "Distância Sensores (cm)")
    for i, o in enumerate(offsets):
        ws.cell(4, 3 + i, o)
    cabecalho = [
        "Distancia_m", "Estaca_Numero", "Load_kgf",
        *[f"D{i}_um" for i in range(9)],
        "Air_C", "Pave_C", "Emod_MPa", "Time",
    ]
    for c, nome in enumerate(cabecalho, start=1):
        ws.cell(6, c, nome)
    for k in range(linhas):
        valores = [k * 20.0, k, 4000 + k, *[400 - 40 * i for i in range(9)],
                   20, 25, 300, "07:5%d:00" % k]
        for c, v in enumerate(valores, start=1):
            ws.cell(7 + k, c, v)
    p = tmp_path / "kuab.xlsx"
    wb.save(p)
    return p


def test_sniffing_reconhece_kuab_e_nao_confunde_solocap(tmp_path):
    assert e_planilha_kuab(_kuab_sintetico(tmp_path)) is True
    assert e_planilha_kuab(RAIZ / "templates" / "Modelo.xlsx") is False
    assert e_planilha_kuab(tmp_path / "nao_existe.xlsx") is False


def test_mapeamento_de_sensores_desloca_em_um(tmp_path):
    """D0_um é o geofone central -> D1 do mapa de Rocha. Sem 10º geofone."""
    d = ler_kuab_xlsx(_kuab_sintetico(tmp_path))
    linha = d.tabela.iloc[0]
    assert linha["D1"] == 400  # D0_um
    assert linha["D2"] == 360  # D1_um
    assert linha["D9"] == 80   # D8_um
    assert linha["D10"] != linha["D10"]  # NaN: KUAB não tem geofone a 210 cm


def test_unidade_micrometro_declarada_pelas_colunas(tmp_path):
    d = ler_kuab_xlsx(_kuab_sintetico(tmp_path))
    assert d.unidade_deflexao is UnidadeDeflexao.MICROMETRO
    assert d.metadados.unidade == "µm"


def test_data_do_cabecalho_combina_com_a_hora_da_linha(tmp_path):
    d = ler_kuab_xlsx(_kuab_sintetico(tmp_path))
    assert d.tabela["Data e Hora"].iloc[0] == "10/03/2018 07:50:00"


def test_tabela_tem_o_mesmo_formato_do_solocap(tmp_path):
    """Downstream (analytics, segmentação, export) não distingue a origem."""
    d = ler_kuab_xlsx(_kuab_sintetico(tmp_path))
    for col in COLUNAS_TABELA:
        assert col in d.tabela.columns, col
    # Extras específicos do KUAB.
    assert "Estaca – Número" in d.tabela.columns
    assert "Emod_MPa" in d.tabela.columns


def test_offsets_divergentes_viram_aviso(tmp_path):
    """Outra geometria de geofones invalidaria o mapeamento — precisa avisar."""
    p = _kuab_sintetico(tmp_path, offsets=(0, 25, 40, 50, 70, 100, 130, 160, 200))
    d = ler_kuab_xlsx(p)
    assert any("Distâncias dos sensores" in a for a in d.avisos)


def test_planilha_sem_d0_um_e_rejeitada(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, "qualquer coisa")
    p = tmp_path / "vazia.xlsx"
    wb.save(p)
    with pytest.raises(ValueError, match="D0_um"):
        ler_kuab_xlsx(p)


@pytest.mark.skipif(not KUAB.exists(), reason="arquivo KUAB ausente")
def test_arquivo_real_do_machado():
    d = carregar(KUAB)  # despacho por conteúdo, não por extensão
    assert len(d.tabela) == 98
    assert d.unidade_deflexao is UnidadeDeflexao.MICROMETRO
    assert d.metadados.cliente == "UNIVERSIDADE FEDERAL DE JUIZ DE FORA"
    assert d.metadados.secao() == "CAMPUS UFJF"
    assert d.metadados.data == "10/03/2018"
    assert not d.avisos  # offsets batem com o mapa de sensores

    linha = d.tabela.iloc[0]
    assert linha["Metros"] == _PRIMEIRA["metros"]
    assert linha["Target Load (Kgf)"] == _PRIMEIRA["carga"]
    assert linha["Estaca – Número"] == _PRIMEIRA["estaca"]
    for i, esperado in enumerate(_PRIMEIRA["d"], start=1):
        assert linha[f"D{i}"] == esperado, f"D{i}"
    assert linha["Data e Hora"] == "10/03/2018 07:50:16"


@pytest.mark.skipif(
    not (KUAB.exists() and SOLOCAP.exists()), reason="arquivos ausentes"
)
def test_solocap_continua_indo_para_o_leitor_solocap():
    d = carregar(SOLOCAP)
    assert len(d.tabela) == 105
    assert d.unidade_deflexao is UnidadeDeflexao.DMM_001


@pytest.mark.skipif(not KUAB.exists(), reason="arquivo KUAB ausente")
def test_fluxo_completo_do_kuab():
    """µm na entrada -> 0,01 mm na planilha padronizada -> µm no CSV BackMeDiNa."""
    from backmedina.convert.backmedina_csv import exportar_csv_backmedina
    from backmedina.convert.standardize import exportar_xlsx_bytes
    from backmedina.model.schema import BACKMEDINA_HEADER

    d = carregar(KUAB)
    assert exportar_xlsx_bytes(d)[:2] == b"PK"

    linhas = exportar_csv_backmedina(d, faixa=1, trilha=2).splitlines()
    col = {n: i for i, n in enumerate(BACKMEDINA_HEADER)}
    assert linhas[1].split(";")[1] == "CAMPUS UFJF"
    pr = linhas[4].split(";")
    assert pr[col["d0"]] == "410"        # fonte já em µm -> sem ×10
    assert pr[col["d180"]] == "26"
    assert pr[col["d210"]] == "0"        # KUAB não tem o 10º geofone
    assert pr[col["Estaca – Número"]] == "0"   # inteiro, não "0.0"
    assert pr[col["Carga"]] == "4097"
    assert pr[col["Data de Execução"]] == "10/03/2018"
