"""Diagnóstico e rascunho .xlsx para PDFs que nenhum leitor reconhece."""

import io
from pathlib import Path

import openpyxl
import pytest

from backmedina.convert.rascunho_pdf import rascunho_xlsx_bytes
from backmedina.io.pdf_fwd import diagnosticar, extrair_texto_layout
from backmedina.io.xlsx_solocap import ler_solocap_xlsx
from backmedina.model.schema import COLUNAS_TABELA, METADADOS_CHAVES
from backmedina.model.units import detectar_unidade_explicita

RAIZ = Path(__file__).resolve().parents[1]
PDF_KUAB = RAIZ / "z_docs" / "lwd" / "machado" / "00_tabela_FWD_machado_2019.pdf"

_TEXTO_KUAB = """                 KUAB                    UNIVERSIDADE FEDERAL
Data                                          10/03/2018
  Distancia_m  Estaca_Numero Load_kgf D0_um D1_um Air_C Pave_C Time
            0              0     4097   410   294    19     25 07:50:16
           20              1     4098   421   274    19     26 07:53:43
"""


def test_diagnostico_aponta_a_ancora_e_a_contagem_de_campos():
    pistas = " ".join(diagnosticar(_TEXTO_KUAB))
    assert "data (dd/mm/aaaa) + hora" in pistas   # âncora ausente
    assert "8 campos" in pistas                   # contagem real das linhas
    assert "18 campos numéricos" in pistas        # o que o SOLOCAP espera


def test_diagnostico_de_pdf_sem_texto():
    assert "não tem texto extraível" in diagnosticar("")[0]
    assert "não tem texto extraível" in diagnosticar("   \n \n")[0]


def test_diagnostico_quando_nenhuma_linha_comeca_com_numero():
    pistas = " ".join(diagnosticar("cabeçalho qualquer\noutra linha de texto"))
    assert "Nenhuma linha começa com um número" in pistas


def _abrir(b: bytes):
    return openpyxl.load_workbook(io.BytesIO(b))


def test_rascunho_tem_esqueleto_e_texto_do_pdf():
    wb = _abrir(rascunho_xlsx_bytes(_TEXTO_KUAB, origem="relatorio.pdf"))
    assert wb.sheetnames == ["Tabela", "Texto_do_PDF"]

    ws = wb["Tabela"]
    assert "RASCUNHO" in str(ws.cell(1, 1).value)
    # Chaves de metadados presentes, valores em branco.
    chaves = [ws.cell(r, 1).value for r in range(3, 3 + len(METADADOS_CHAVES))]
    assert chaves == list(METADADOS_CHAVES)
    assert all(ws.cell(r, 2).value is None for r in range(3, 3 + len(METADADOS_CHAVES)))
    # Cabeçalho das colunas na linha 13, como o layout SOLOCAP.
    assert [ws.cell(13, c).value for c in range(1, len(COLUNAS_TABELA) + 1)] == list(
        COLUNAS_TABELA
    )

    wt = wb["Texto_do_PDF"]
    assert wt.cell(3, 1).value == "KUAB"          # 1ª linha do texto, quebrada
    assert wt.cell(6, 1).value == "0"             # 1ª linha de dados
    assert wt.cell(6, 4).value == "410"


def test_rascunho_nao_declara_unidade(tmp_path):
    """A armadilha do erro de 10x: um default aqui contaminaria tudo a jusante."""
    p = tmp_path / "rascunho.xlsx"
    p.write_bytes(rascunho_xlsx_bytes(_TEXTO_KUAB))
    ws = openpyxl.load_workbook(p, data_only=True)["Tabela"]
    linha_unidade = 3 + list(METADADOS_CHAVES).index("UNIDADE DAS LEITURAS")
    assert ws.cell(linha_unidade, 1).value == "UNIDADE DAS LEITURAS"
    assert ws.cell(linha_unidade, 2).value is None

    # Relido pelo App: "não declarada" -> a UI pede confirmação ao usuário.
    dados = ler_solocap_xlsx(p)
    assert dados.metadados.unidade == ""
    assert detectar_unidade_explicita(dados.metadados.unidade) is None


def test_rascunho_vazio_continua_sendo_import_bloqueado(tmp_path):
    """O rascunho não vem com dados: reimportá-lo sem preencher é import vazio."""
    p = tmp_path / "rascunho.xlsx"
    p.write_bytes(rascunho_xlsx_bytes(_TEXTO_KUAB))
    assert len(ler_solocap_xlsx(p).tabela) == 0


@pytest.mark.skipif(not PDF_KUAB.exists(), reason="PDF do Machado ausente")
def test_pdf_real_do_machado_gera_diagnostico_util():
    texto = extrair_texto_layout(PDF_KUAB)
    pistas = " ".join(diagnosticar(texto))
    assert "16 campos" in pistas   # 3 + 9 sensores + 2 temps + Emod + Time
    assert "98 linha(s)" in pistas
    wb = _abrir(rascunho_xlsx_bytes(texto, origem=PDF_KUAB.name))
    assert wb["Texto_do_PDF"].max_row > 100
