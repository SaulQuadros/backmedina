import pandas as pd

from backmedina.analytics.pbd_indices import calcular_indices
from backmedina.convert.backmedina_csv import exportar_csv_backmedina
from backmedina.model.schema import BACKMEDINA_HEADER, MetadadosLevantamento, DadosFWD
from backmedina.model.units import (
    UnidadeDeflexao,
    detectar_unidade,
    detectar_unidade_explicita,
    fator_conversao,
)


def test_detectar_unidade():
    assert detectar_unidade("UNIDADE DAS LEITURAS (x10⁻² mm)") is UnidadeDeflexao.DMM_001
    assert detectar_unidade("Deflexões em µm") is UnidadeDeflexao.MICROMETRO
    assert detectar_unidade("micrômetro") is UnidadeDeflexao.MICROMETRO
    # Sem marcador reconhecível -> default.
    assert detectar_unidade("texto qualquer") is UnidadeDeflexao.DMM_001


def test_detectar_unidade_explicita_distingue_ausente_de_declarado():
    """A UI precisa separar 'declarou 0,01 mm' de 'não declarou nada'."""
    assert detectar_unidade_explicita("(x10⁻² mm)") is UnidadeDeflexao.DMM_001
    assert detectar_unidade_explicita("µm") is UnidadeDeflexao.MICROMETRO
    assert detectar_unidade_explicita("texto qualquer") is None
    assert detectar_unidade_explicita("") is None
    assert detectar_unidade_explicita(None) is None


def test_grafia_ascii_um_do_kuab():
    """KUAB nomeia as colunas D0_um/D1_um — sem casar dentro de outras palavras."""
    assert detectar_unidade_explicita("D0_um D1_um D2_um") is UnidadeDeflexao.MICROMETRO
    assert detectar_unidade_explicita("Leituras (um)") is UnidadeDeflexao.MICROMETRO
    # Armadilhas: 'um' colado a letras não é unidade.
    for texto in ("Estaca_Numero", "resumo do trecho", "volume", "algum ponto",
                  "alumínio", "UNIDADE DAS LEITURAS"):
        assert detectar_unidade_explicita(texto) is None, texto
    # Marcador forte vence o fraco quando os dois aparecem.
    assert detectar_unidade_explicita(
        "um levantamento com leituras em x10⁻² mm"
    ) is UnidadeDeflexao.DMM_001


def test_planilha_padronizada_converte_fonte_em_micrometro(tmp_path):
    """Round-trip que travava o erro de 10x: µm -> planilha -> releitura.

    A planilha declara `x10⁻² mm` na linha 10; antes gravava os valores crus da
    fonte, então reimportá-la multiplicava tudo por 10.
    """
    import openpyxl

    from backmedina.convert.standardize import exportar_xlsx_bytes
    from backmedina.io.xlsx_solocap import ler_solocap_xlsx

    dados = _dados(UnidadeDeflexao.MICROMETRO)  # D1 = 44 µm
    p = tmp_path / "Tabela_padronizada.xlsx"
    p.write_bytes(exportar_xlsx_bytes(dados))

    ws = openpyxl.load_workbook(p, data_only=True).active
    assert "x10⁻² mm" in str(ws.cell(10, 2).value)
    assert ws.cell(14, 4).value == 4.4  # 44 µm -> 4,4 (0,01 mm)

    relido = ler_solocap_xlsx(p)
    assert relido.unidade_deflexao is UnidadeDeflexao.DMM_001
    assert relido.tabela["D1"].iloc[0] == 4.4  # mesma deflexão física
    # A fonte não é alterada pela exportação.
    assert dados.tabela["D1"].iloc[0] == 44


def test_planilha_padronizada_preserva_fonte_em_0p01mm(tmp_path):
    """Fonte já em 0,01 mm (SOLOCAP): nenhuma conversão, valores intactos."""
    import openpyxl

    from backmedina.convert.standardize import exportar_xlsx_bytes

    p = tmp_path / "Tabela_padronizada.xlsx"
    p.write_bytes(exportar_xlsx_bytes(_dados(UnidadeDeflexao.DMM_001)))
    ws = openpyxl.load_workbook(p, data_only=True).active
    assert ws.cell(14, 4).value == 44
    assert ws.cell(14, 13).value == 2  # D10 também intacto


def test_metadados_nascem_sem_unidade_declarada():
    """Default vazio = 'o arquivo não declarou'.

    Com um default de 0,01 mm, todo arquivo parecia ter se identificado e uma
    fonte em µm passava valendo 10x sem que a UI pudesse alertar.
    """
    assert MetadadosLevantamento().unidade == ""
    assert detectar_unidade_explicita(MetadadosLevantamento().unidade) is None


def test_planilha_sem_linha_de_unidade_nao_herda_default(tmp_path):
    """Sem 'UNIDADE DAS LEITURAS' na planilha, o metadado fica vazio.

    O default do dataclass diz 'x10⁻² mm'; herdá-lo faria a UI tratar um arquivo
    silencioso como se ele tivesse declarado a unidade.
    """
    from openpyxl import Workbook

    from backmedina.io.xlsx_solocap import ler_solocap_xlsx
    from backmedina.model.schema import COLUNAS_TABELA

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela"
    ws.cell(3, 1, "DATA"); ws.cell(3, 2, "10/03/2018")  # sem a linha de unidade
    for c, nome in enumerate(COLUNAS_TABELA, start=1):
        ws.cell(13, c, nome)
    ws.cell(14, 1, 0); ws.cell(14, 4, 44)
    p = tmp_path / "sem_unidade.xlsx"
    wb.save(p)

    dados = ler_solocap_xlsx(p)
    assert dados.metadados.unidade == ""
    assert detectar_unidade_explicita(dados.metadados.unidade) is None
    # O fluxo segue com o padrão — mas agora a UI sabe que foi presumido.
    assert dados.unidade_deflexao is UnidadeDeflexao.DMM_001


def test_fatores():
    assert UnidadeDeflexao.DMM_001.fator_para_micrometro == 10.0
    assert UnidadeDeflexao.MICROMETRO.fator_para_micrometro == 1.0
    assert fator_conversao(UnidadeDeflexao.DMM_001, UnidadeDeflexao.MICROMETRO) == 10.0
    assert fator_conversao(UnidadeDeflexao.MICROMETRO, UnidadeDeflexao.DMM_001) == 0.1


def _dados(unidade):
    tab = pd.DataFrame(
        [{"Metros": 0.0, "Target Load (Kgf)": 4000,
          "D1": 44, "D2": 29, "D3": 21, "D4": 13, "D5": 9,
          "D6": 6, "D7": 4, "D8": 3, "D9": 3, "D10": 2,
          "Data e Hora": "11/09/2023 13:30"}]
    )
    return DadosFWD(MetadadosLevantamento(), tab, unidade_deflexao=unidade)


def test_csv_converte_0p01mm_para_micrometro():
    csv = exportar_csv_backmedina(_dados(UnidadeDeflexao.DMM_001))
    col = {n: i for i, n in enumerate(BACKMEDINA_HEADER)}
    linha = csv.splitlines()[4].split(";")
    assert linha[col["d0"]] == "440"   # 44 * 10
    assert linha[col["d180"]] == "30"  # 3 * 10


def test_csv_nao_converte_quando_ja_micrometro():
    csv = exportar_csv_backmedina(_dados(UnidadeDeflexao.MICROMETRO))
    col = {n: i for i, n in enumerate(BACKMEDINA_HEADER)}
    linha = csv.splitlines()[4].split(";")
    assert linha[col["d0"]] == "44"  # sem conversão
    assert linha[col["d180"]] == "3"


def test_indices_invariantes_a_unidade():
    # Mesma bacia física expressa em 0,01 mm e em µm -> índices idênticos.
    base = {"D1": 44, "D2": 29, "D3": 21, "D4": 13, "D5": 9,
            "D6": 6, "D7": 4, "D8": 3, "D9": 3, "D10": 2}
    df_dmm = pd.DataFrame([base])
    df_um = pd.DataFrame([{k: v * 10 for k, v in base.items()}])
    a = calcular_indices(df_dmm, unidade=UnidadeDeflexao.DMM_001).iloc[0]
    b = calcular_indices(df_um, unidade=UnidadeDeflexao.MICROMETRO).iloc[0]
    for col in ("Rc", "AREA", "SCI", "BDI", "BCI", "CF", "S"):
        assert abs(a[col] - b[col]) < 1e-6, col


def test_csvs_por_segmento_zip():
    import io as _io
    import zipfile
    from backmedina.convert.backmedina_csv import (
        csvs_por_segmento, zip_de_arquivos,
    )
    from backmedina.io.xlsx_solocap import ler_solocap_xlsx
    from backmedina.segmentation.aashto_cumdiff import segmentar
    from pathlib import Path

    xlsx = Path(__file__).resolve().parents[1] / "z_docs" / "lwd" / "solocap" / "2-UFJF-VIA_LOCAL_FX1-FWD.xlsx"
    if not xlsx.exists():
        import pytest; pytest.skip("arquivo UFJF ausente")

    dados = ler_solocap_xlsx(xlsx)
    df_seg, segs = segmentar(dados.tabela, unidade=dados.unidade_deflexao)
    arquivos = csvs_por_segmento(dados, df_seg)

    # Um CSV por segmento, nomeados seg_01, seg_02, ...
    assert len(arquivos) == len(segs)
    assert "seg_01.csv" in arquivos
    esperado = {f"seg_{i:02d}.csv" for i in range(1, len(segs) + 1)}
    assert set(arquivos) == esperado

    # Cada CSV: cabeçalho correto, SEÇÃO seg_NN, deflexões em µm (D1*10).
    csv1 = arquivos["seg_01.csv"].splitlines()
    assert csv1[0].split(";")[0] == "BACKMEDINA"
    assert csv1[1].split(";")[:2] == ["SEÇÃO:", "seg_01"]
    assert csv1[2].split(";")[:2] == ["RAIO (cm):", "15"]
    d0_primeira = int(csv1[4].split(";")[8])   # coluna d0
    d0_bruto = int(dados.tabela["D1"].iloc[0])
    assert d0_primeira == d0_bruto * 10

    # Nº de estações no seg_01 == nº de linhas de dados do CSV (linhas - 4).
    assert len(csv1) - 4 == segs[0].n_pontos

    # ZIP contém todos os arquivos.
    zb = zip_de_arquivos(arquivos)
    with zipfile.ZipFile(_io.BytesIO(zb)) as zf:
        assert set(zf.namelist()) == esperado


def test_resumo_seg_xlsx_e_zip_com_resumo():
    import io as _io
    import zipfile
    from pathlib import Path
    from openpyxl import load_workbook
    from backmedina.convert.backmedina_csv import csvs_por_segmento, zip_de_arquivos
    from backmedina.convert.resumo_seg import resumo_seg_xlsx_bytes
    from backmedina.io.xlsx_solocap import ler_solocap_xlsx
    from backmedina.segmentation.aashto_cumdiff import segmentar

    xlsx = Path(__file__).resolve().parents[1] / "z_docs" / "lwd" / "solocap" / "2-UFJF-VIA_LOCAL_FX1-FWD.xlsx"
    if not xlsx.exists():
        import pytest; pytest.skip("arquivo UFJF ausente")

    dados = ler_solocap_xlsx(xlsx)
    df_seg, segs = segmentar(dados.tabela, unidade=dados.unidade_deflexao)

    # Resumo_Seg.xlsx válido, com uma linha por segmento.
    data = resumo_seg_xlsx_bytes(segs)
    assert data[:2] == b"PK"
    wb = load_workbook(_io.BytesIO(data))
    ws = wb["Resumo_Seg"]
    header = [c.value for c in ws[4]]
    assert header == ["Segmento", "Estacas (m)", "Comprimento (m)", "Dm", "σ", "Dc"]
    # primeira linha de dados = seg_01
    assert ws.cell(5, 1).value == "seg_01"

    # ZIP contém os CSVs + o Resumo_Seg.xlsx.
    arquivos = csvs_por_segmento(dados, df_seg)
    arquivos["Resumo_Seg.xlsx"] = data
    zb = zip_de_arquivos(arquivos)
    with zipfile.ZipFile(_io.BytesIO(zb)) as zf:
        nomes = set(zf.namelist())
    assert "Resumo_Seg.xlsx" in nomes
    assert "seg_01.csv" in nomes
    assert len([n for n in nomes if n.endswith(".csv")]) == len(segs)


def test_glossario_cobre_todas_as_colunas():
    import pandas as pd
    from backmedina.analytics.pbd_glossario import GLOSSARIO_PBD
    from backmedina.analytics.pbd_indices import indices_com_contexto
    df = pd.DataFrame([{"D1":44,"D2":29,"D3":21,"D4":13,"D5":9,"D6":6,"D7":4,"D8":3,"D9":3,"D10":2,"Metros":0.0}])
    colunas = set(indices_com_contexto(df).columns)  # Metros, Rc, AREA, SCI, BDI, BCI, CF, S
    siglas = {g["sigla"] for g in GLOSSARIO_PBD}
    assert colunas <= siglas, f"faltam no glossário: {colunas - siglas}"
    # cada item tem os campos essenciais preenchidos
    for g in GLOSSARIO_PBD:
        for campo in ("nome","unidade","formula","o_que_e","para_que_serve"):
            assert g[campo], f"{g['sigla']} sem {campo}"
